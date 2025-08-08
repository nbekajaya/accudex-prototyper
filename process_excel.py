import datetime
import os
import re
import openpyxl

import pandas as pd

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.alignment import Alignment
from toolbox import Toolbox
    
def get_custom_landmark_info(custom_landmark) -> dict:
    def get_values(space:str):
        return {f'{name.title()}_{space.title()}_{el_name}':value 
                for el_name, value 
                in zip('xyz',getattr(custom_landmark,space))}
    name:str = custom_landmark.name
    return get_values('screen')|get_values('world')

def get_measurement_info(measurement_dict:dict) -> dict:
    if 'check_index' not in measurement_dict['params']:
        return None
    params:dict = measurement_dict['params']
    function_name:str = measurement_dict['function_name']
    result_index:int = params['check_index']
    default_name:str = '-'.join(params['default_names'])
    return {'name' : measurement_dict.get('name', default_name)\
             + f'_{function_name}'\
             + f'_{result_index}'}\
            | params

def apply_style(s:pd.DataFrame, parameters:dict=None) -> pd.DataFrame:
    styler = pd.DataFrame('',s.index,s.columns)
    style_list = (
        'background-color:#f78686',
        'background-color:#f7d586',
        'background-color:#86f7a6',
        
    )
    for param_idx, param in enumerate(parameters):
        column_idx:int  = param_idx
        comparison_series = [style_list[Toolbox.value_discriminator(x, param)] 
                            for x 
                            in s.iloc[:,column_idx]]
        styler.iloc[:, param_idx] = comparison_series
    return styler

def apply_style(s:pd.DataFrame, parameters:dict=None) -> pd.DataFrame:
    styler = pd.DataFrame('',s.index,s.columns)
    style_list = (
        'background-color:#f78686',
        'background-color:#f7d586',
        'background-color:#86f7a6',
        
    )
    for param_idx, param in enumerate(parameters):
        column_idx:int  = param_idx
        comparison_series = [style_list[Toolbox.value_discriminator(x, param)] 
                            for x 
                            in s.iloc[:,column_idx]]
        styler.iloc[:, param_idx] = comparison_series
    return styler

def recording_info_to_pd_dataframe(recording:list[int,list[list],list[dict]]) -> pd.DataFrame:
    """Converts the recorded data from the loop to a dataframe

    Args:
        recording (list[int,list[list],list[dict]]): A recording which is a list of lists containing timestamp
                                                     landmark_list, and measurement dictionary

    Returns:
        pd.DataFrame: Dataframe with style applied
    """
    recording = [frame for frame in recording if len(frame[1])>0] # Make sure all data are safe

    # ENSURE NO DUPLICATES
    timestamps = [frame_info[0] for frame_info in recording]
    duplicated_idx = [t_idx for t_idx, timestamp in enumerate(timestamps) 
                        if timestamps.count(timestamp)>1]
    duplicated_idx = [duplicated_idx[i:i+2][-1] for i in range(0,len(duplicated_idx),2)]
    recording = [frame for idx, frame in enumerate(recording) if idx not in duplicated_idx]

    # GET TIMESTAMPS
    recorded_timestamps = {'timestamp_ms':[(frame_info[0]-recording[0][0])/1000 for frame_info in recording]}
    timestamps_df = pd.DataFrame(recorded_timestamps)

    # GET LANDMARK INFOMATION
    recorded_landmarks = [[get_custom_landmark_info(cl) 
                           for cl in sorted(frame_info[1][0],key=lambda x:x.idx)] 
                          for frame_info in recording]
    landmarks_keys = sum([list(d.keys()) for d in recorded_landmarks[0]],[])
    landmarks_vals = list(zip(*[sum([list(d.values()) for d in frame],[]) 
                                for frame in recorded_landmarks]))
    landmarks_data = {k:v for k,v in zip(landmarks_keys,landmarks_vals)}
    landmarks_df = pd.DataFrame(landmarks_data)

    # GET MEASUREMENTS
    recorded_measurements = [frame_info[2][0] for frame_info in recording]
    measurements_info = [m_info
                        for measurement in recorded_measurements[0]
                        if (m_info:=get_measurement_info(measurement)) is not None]
    measurements_keys = [measurement['name'] for measurement in measurements_info]
    measurements_vals = list(zip(*[[measurement['result'][measurement['params']['check_index']]
                            for measurement in frame]
                            for frame in recorded_measurements]))
    measurements_data = {k:v for k,v in zip(measurements_keys,measurements_vals)}
    measurements_df = pd.DataFrame(measurements_data)
    
    df = pd.concat([timestamps_df,measurements_df,landmarks_df],axis=1).set_index('timestamp_ms')
    try:
        applied_style = df.style.apply(apply_style, parameters = measurements_info, axis = None)
    except:
        applied_style = None

    return applied_style

def write_to_excel(dataframe:pd.DataFrame, file_name:str, directory:str='.'):
    """Generates an excel file from the given recording dataframe using openpyxl

    Args:
        dataframe (pd.DataFrame): A pandas dataframe of the recording info
        file_name (str): The name of the file; This is combined with the current date during export
        directory (str, optional): The directory to write the file to. Defaults to '.'.
    """
    try:
        assert os.path.isdir(os.path.abspath(directory))
    except AssertionError:
        os.mkdir(directory:=os.path.abspath(directory))

    # SET EXPORT NAME
    now = datetime.datetime.now()
    now = f'_{now.year}{now.month:02d}{now.day:02d}-{now.hour:02d}{now.minute:02d}'
    file_name = file_name+now
    tmp_name = file_name
    addition = 1
    while os.path.isfile(os.path.join(directory,tmp_name+'.xlsx')):
        tmp_name = file_name+f'({addition})'
        addition += 1
    file_name = os.path.join(directory,tmp_name+'.xlsx')

    # EXPORT FROM PANDAS
    dataframe.to_excel(file_name)

    # FIX STUFF WITH OPENPYXL
    wb = openpyxl.load_workbook(file_name)
    ws = wb.worksheets[0]
    max_column = column_index_from_string(
                    re.search(r'\D+',ws.dimensions.split(':')[1]).group())
    ws.column_dimensions['A'].width = 15
    for header in ws.iter_cols(min_col=1,max_row=1):
        for cell in header:
            cell.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    wrap_text=True)
    for col_id in range(2,max_column+1):
        ws.column_dimensions[get_column_letter(col_id)].width = 20
    ws.freeze_panes = ws['B1']
    wb.save(file_name)

if __name__ == '__main__':
    pass