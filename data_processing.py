import datetime
import os
import re
import openpyxl

import numpy as np
import pandas as pd
from models import CustomLandmark

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.alignment import Alignment
from toolbox import Toolbox
    
def get_custom_landmark_info(custom_landmark) -> dict:
    def get_values(space:str):
        return {f'{name.title()}_{space.title()}_{el_name}':value 
                for el_name, value 
                in zip('xyz',getattr(custom_landmark,space))}
    name:str = custom_landmark.name
    return get_values('screen')|get_values('world')

def get_measurement_info(measurement_dict:dict) -> dict:
    if 'check_index' not in measurement_dict['params']:
        return None
    params:dict = measurement_dict['params']
    function_name:str = measurement_dict['function_name']
    result_index:int = params['check_index']
    default_name:str = '-'.join(params['default_names'])
    return {'name' : measurement_dict.get('name', default_name)\
             + f'_{function_name}'\
             + f'_{result_index}'}\
            | params

def apply_style(s:pd.DataFrame, parameters:dict=None) -> pd.DataFrame:
    styler = pd.DataFrame('',s.index,s.columns)
    style_list = (
        'background-color:#f78686',
        'background-color:#f7d586',
        'background-color:#86f7a6',
        
    )
    for param_idx, param in enumerate(parameters):
        column_idx:int  = param_idx
        comparison_series = [style_list[Toolbox.value_discriminator(x, param)] 
                            for x 
                            in s.iloc[:,column_idx]]
        styler.iloc[:, param_idx] = comparison_series
    return styler

def recording_info_to_pd_dataframe(recording: dict[int,list[dict[str,int]]]) -> pd.DataFrame:
    """Converts the recorded data from the loop to a dataframe

    Args:
        recording (dict[int,list[dict[str,int]]]): A recording which is a dictionary 
        with timestamp as keys and a list of dictionaries with keys: 'landmark','measure'

    Returns:
        pd.DataFrame: Dataframe with style applied
    """
    # recording = [frame for frame in recording if len(frame[1])>0] # Make sure all data are safe

    # ENSURE NO DUPLICATES
    # timestamps = [frame_info[0] for frame_info in recording]
    # duplicated_idx = [t_idx for t_idx, timestamp in enumerate(timestamps) 
    #                     if timestamps.count(timestamp)>1]
    # duplicated_idx = [duplicated_idx[i:i+2][-1] for i in range(0,len(duplicated_idx),2)]
    # recording = [frame for idx, frame in enumerate(recording) if idx not in duplicated_idx]

    # GET TIMESTAMPS
    first_timestamp = list(recording.keys())
    recorded_timestamps = {'timestamp_ms':[(frame_info[0]-recording[0][0]) for frame_info in recording]}
    timestamps_df = pd.DataFrame(recorded_timestamps)

    # GET LANDMARK INFOMATION
    recorded_landmarks = [[get_custom_landmark_info(cl) 
                           for cl in sorted(frame_info[1][0],key=lambda x:x.idx)] 
                          for frame_info in recording]
    landmarks_keys = sum([list(d.keys()) for d in recorded_landmarks[0]],[])
    landmarks_vals = list(zip(*[sum([list(d.values()) for d in frame],[]) 
                                for frame in recorded_landmarks]))
    landmarks_data = {k:v for k,v in zip(landmarks_keys,landmarks_vals)}
    landmarks_df = pd.DataFrame(landmarks_data)

    # GET MEASUREMENTS
    recorded_measurements = [measurement for frame_info in recording
                             if len((measurement:=frame_info[2][0]))>0] + [[]]
    measurements_info = [m_info
                        for measurement in recorded_measurements[0]
                        if (m_info:=get_measurement_info(measurement)) is not None]
    measurements_keys = [measurement['name'] for measurement in measurements_info]
    measurements_vals = list(zip(*[[measurement['result'][measurement['params']['check_index']]
                            for measurement in frame]
                            for frame in recorded_measurements]))
    measurements_data = {k:v for k,v in zip(measurements_keys,measurements_vals)}
    measurements_df = pd.DataFrame(measurements_data)
    
    df = pd.concat([timestamps_df,measurements_df,landmarks_df],axis=1).set_index('timestamp_ms')
    try:
        applied_style = df.style.apply(apply_style, parameters = measurements_info, axis = None)
    except:
        applied_style = None

    return applied_style

def write_to_excel(dataframe:pd.DataFrame, file_name:str, directory:str='.'):
    """Generates an excel file from the given recording dataframe using openpyxl

    Args:
        dataframe (pd.DataFrame): A pandas dataframe of the recording info
        file_name (str): The name of the file; This is combined with the current date during export
        directory (str, optional): The directory to write the file to. Defaults to '.'.
    """
    try:
        assert os.path.isdir(os.path.abspath(directory))
    except AssertionError:
        os.mkdir(directory:=os.path.abspath(directory))

    # SET EXPORT NAME
    now = datetime.datetime.now()
    now = f'_{now.year}{now.month:02d}{now.day:02d}-{now.hour:02d}{now.minute:02d}'
    file_name = file_name+now
    tmp_name = file_name
    addition = 1
    while os.path.isfile(os.path.join(directory,tmp_name+'.xlsx')):
        tmp_name = file_name+f'({addition})'
        addition += 1
    file_name = os.path.join(directory,tmp_name+'.xlsx')

    # EXPORT FROM PANDAS
    dataframe.to_excel(file_name)

    # FIX STUFF WITH OPENPYXL
    wb = openpyxl.load_workbook(file_name)
    ws = wb.worksheets[0]
    max_column = column_index_from_string(
                    re.search(r'\D+',ws.dimensions.split(':')[1]).group())
    ws.column_dimensions['A'].width = 15
    for header in ws.iter_cols(min_col=1,max_row=1):
        for cell in header:
            cell.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    wrap_text=True)
    for col_id in range(2,max_column+1):
        ws.column_dimensions[get_column_letter(col_id)].width = 20
    ws.freeze_panes = ws['B1']
    wb.save(file_name)

def process_com_from_recording(recording:list[int,list[list],dict]) -> dict[dict,dict]:
    """Calculates biomechanical statistics from recorded data for center of mass.
    Center of mass is identified by the landmark with index 5000 in every landmark group

    Args:
        recording (list[int,list[list],dict]): Recording data

    Raises:
        ValueError: If index 5000 is not the center of mass

    Returns:
        dict[dict,dict]: The function returns a dictionary with two keys for two dictionaries:
        - data:
            - position_offsets: offsets of CoM position from average throughout recording in cm
            - position_distances: distances of CoM position from average throughout recording in cm
            - travel_offsets: offsets of CoM position from previous frame's position throughout recording in cm
            - travel_distances: distances of CoM position from previous frame's position throughout recording in cm
            - travel_vels: velocities of CoM from previous frame's position throughout recording in cm/s
            - travel_speeds: speeds of CoM from previous frame's position throughout recording in cm/s
        - info:
            - time_per_frame: average delta time of frames from previous frame in s
            - position_offset_std: standard deviation of CoM position from average in cm
            - position_total_distance: total of CoM distance from average position in cm
            - position_total_displacement: total of CoM displacement from average position in cm
            - travel_total_distance: total of CoM travel distance (one frame to the next) in cm
            - travel_total_displacement: total of CoM travel displacement (one frame to the next) in cm
            - position_avg_vel: average of CoM velocity from average position in cm/s
            - position_avg_speed: average of CoM speed from average position in cm/s
            - travel_avg_vel: average of CoM travel velocity (one frame to the next) in cm/s
            - travel_avg_speed: average of CoM travel speed (one frame to the next) in cm/s
    """
    template:list[CustomLandmark] = recording[0][1][0] # LANDMARK LIST
    index_map = {landmark.idx:landmark.name for landmark in template}
    result_data, result_info = dict(), dict()

    try:
        assert index_map[5000] == 'center of mass'
    except AssertionError:
        raise ValueError("Center of Mass not in expected index 5000")

    all_timestamps = [record[0] for record in recording]
    all_frame_durations = [
        (all_timestamps[i]-all_timestamps[i-1]) / 1000 
        for i in range(1,len(all_timestamps))
    ]
    average_frame_duration = np.mean(all_frame_durations)

    com_data = [
        np.array(record[1][0][-1].world[::2])*100 for record in recording
    ]
    com_avg = np.mean(com_data,0)

    com_displacement = [(com - com_avg) for com in com_data]
    com_distance = [
        np.sqrt(sum([x**2 for x in displacement])) for displacement in com_displacement
    ]
    
    com_std = np.std(com_displacement, 0)

    com_total_distance = sum(com_distance)
    com_total_displacement = sum(com_displacement)
    
    com_travel_displacement = [np.array([0.0,0.0])]+[
        com_displacement[i] - com_displacement[i-1] for i in range(1, len(com_displacement))
    ]
    com_travel_distance = [0.0]+[
        np.sqrt(sum([x**2 for x in displacement])) for displacement in com_travel_displacement
    ]
    
    com_total_travel_displacement = sum(com_travel_displacement)
    com_total_travel_distance = sum(com_travel_distance)
    
    com_travel_velocity = [np.array([0.0,0.0])]+[
        displacement/average_frame_duration for displacement in com_travel_displacement
    ]
    com_travel_speed = [0.0]+[
        distance/average_frame_duration for distance in com_travel_distance
    ]

    # ARRAYS
    result_data["position_offsets"] = com_displacement
    result_data["position_distances"] = com_distance 
    result_data["travel_offsets"] = com_travel_displacement
    result_data["travel_distances"] = com_travel_distance
    result_data["travel_vels"] = com_travel_velocity
    result_data["travel_speeds"] = com_travel_speed
    
    # SINGULAR VALUES
    result_info["time_per_frame"] = average_frame_duration
    result_info["position_offset_std"] = com_std

    result_info["position_total_distance"] = com_total_distance
    result_info["position_total_displacement"] = com_total_displacement
    result_info["travel_total_distance"] = com_total_travel_distance
    result_info["travel_total_displacement"] = com_total_travel_displacement
    
    result_info["position_avg_vel"] = com_total_displacement/(average_frame_duration*len(com_displacement))
    result_info["position_avg_speed"] = com_total_distance/(average_frame_duration*len(com_distance))
    result_info["travel_avg_vel"] = np.mean(com_travel_velocity,0)
    result_info["travel_avg_speed"] = np.mean(com_travel_speed)

    return {'info':result_info,'data':result_data}

if __name__ == '__main__':
    pass